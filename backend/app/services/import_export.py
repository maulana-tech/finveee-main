"""
Finvee Backend - Import/Export Services
CSV and Excel handling for transactions and budgets
"""

import csv
import io
import os
from datetime import datetime
from typing import List, Dict, Any

try:
    import pandas as pd

    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from openpyxl import Workbook

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ImportExporter:
    """Handle CSV and Excel import/export"""

    @staticmethod
    def export_transactions_csv(
        transactions: List, filename: str = "transactions.csv"
    ) -> str:
        """Export transactions to CSV"""
        output = io.StringIO()

        fieldnames = [
            "date",
            "type",
            "amount",
            "category",
            "description",
            "account",
            "tags",
            "fraud_score",
        ]
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()

        for t in transactions:
            writer.writerow(
                {
                    "date": t.date,
                    "type": t.type,
                    "amount": t.amount,
                    "category": t.category,
                    "description": t.description or "",
                    "account": t.account_id,
                    "tags": ",".join(t.tags) if t.tags else "",
                    "fraud_score": t.fraud_score or "",
                }
            )

        return output.getvalue()

    @staticmethod
    def import_transactions_csv(
        csv_content: str, user_id: str, account_id: str
    ) -> Dict[str, Any]:
        """Import transactions from CSV"""
        input_stream = io.StringIO(csv_content)
        reader = csv.DictReader(input_stream)

        imported = []
        errors = []

        for row_num, row in enumerate(reader, start=2):
            try:
                # Map CSV columns to transaction fields
                date = row.get("date", "").strip()
                trans_type = row.get("type", "").strip().lower()
                amount = float(row.get("amount", 0))
                category = row.get("category", "Other").strip()
                description = row.get("description", "").strip()
                tags = row.get("tags", "").strip()

                if not date:
                    errors.append(f"Row {row_num}: Missing date")
                    continue

                if trans_type not in ["income", "expense"]:
                    errors.append(f"Row {row_num}: Invalid type '{trans_type}'")
                    continue

                if amount <= 0:
                    errors.append(f"Row {row_num}: Invalid amount")
                    continue

                imported.append(
                    {
                        "date": date,
                        "type": trans_type,
                        "amount": amount,
                        "category": category,
                        "description": description,
                        "tags": tags.split(",") if tags else [],
                    }
                )

            except (ValueError, KeyError) as e:
                errors.append(f"Row {row_num}: {str(e)}")

        return {
            "success": True,
            "imported": imported,
            "errors": errors,
            "total": len(imported),
        }

    @staticmethod
    def export_budgets_csv(budgets: List, filename: str = "budgets.csv") -> str:
        """Export budgets to CSV"""
        output = io.StringIO()

        fieldnames = ["category", "amount", "period", "start_date", "end_date"]
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()

        for b in budgets:
            writer.writerow(
                {
                    "category": b.category,
                    "amount": b.amount,
                    "period": b.period,
                    "start_date": b.start_date,
                    "end_date": b.end_date,
                }
            )

        return output.getvalue()

    @staticmethod
    def import_budgets_csv(csv_content: str, user_id: str) -> Dict[str, Any]:
        """Import budgets from CSV"""
        input_stream = io.StringIO(csv_content)
        reader = csv.DictReader(input_stream)

        imported = []
        errors = []

        for row_num, row in enumerate(reader, start=2):
            try:
                category = row.get("category", "").strip()
                amount = float(row.get("amount", 0))
                period = row.get("period", "monthly").strip().lower()
                start_date = row.get("start_date", "").strip()
                end_date = row.get("end_date", "").strip()

                if not category:
                    errors.append(f"Row {row_num}: Missing category")
                    continue

                if amount <= 0:
                    errors.append(f"Row {row_num}: Invalid amount")
                    continue

                if period not in ["weekly", "monthly", "yearly"]:
                    errors.append(f"Row {row_num}: Invalid period '{period}'")
                    continue

                imported.append(
                    {
                        "category": category,
                        "amount": amount,
                        "period": period,
                        "start_date": start_date,
                        "end_date": end_date,
                    }
                )

            except (ValueError, KeyError) as e:
                errors.append(f"Row {row_num}: {str(e)}")

        return {
            "success": True,
            "imported": imported,
            "errors": errors,
            "total": len(imported),
        }

    @staticmethod
    def export_accounts_csv(accounts: List, filename: str = "accounts.csv") -> str:
        """Export accounts to CSV"""
        output = io.StringIO()

        fieldnames = ["name", "account_type", "balance", "currency", "is_active"]
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()

        for a in accounts:
            writer.writerow(
                {
                    "name": a.name,
                    "account_type": a.account_type,
                    "balance": a.balance,
                    "currency": a.currency,
                    "is_active": a.is_active,
                }
            )

        return output.getvalue()

    @staticmethod
    def import_accounts_csv(csv_content: str, user_id: str) -> Dict[str, Any]:
        """Import accounts from CSV"""
        input_stream = io.StringIO(csv_content)
        reader = csv.DictReader(input_stream)

        imported = []
        errors = []

        for row_num, row in enumerate(reader, start=2):
            try:
                name = row.get("name", "").strip()
                account_type = row.get("account_type", "checking").strip().lower()
                balance = float(row.get("balance", 0))
                currency = row.get("currency", "USD").strip().upper()

                if not name:
                    errors.append(f"Row {row_num}: Missing name")
                    continue

                if account_type not in ["checking", "savings", "credit", "investment"]:
                    errors.append(
                        f"Row {row_num}: Invalid account type '{account_type}'"
                    )
                    continue

                imported.append(
                    {
                        "name": name,
                        "account_type": account_type,
                        "balance": balance,
                        "currency": currency,
                    }
                )

            except (ValueError, KeyError) as e:
                errors.append(f"Row {row_num}: {str(e)}")

        return {
            "success": True,
            "imported": imported,
            "errors": errors,
            "total": len(imported),
        }

    @staticmethod
    def generate_sample_csv(csv_type: str) -> str:
        """Generate sample CSV for user reference"""
        if csv_type == "transactions":
            return """date,type,amount,category,description,tags
2024-01-15,expense,25.50,Food & Dining,Lunch at restaurant,food,lunch
2024-01-16,income,5000.00,Salary,Monthly salary,
2024-01-17,expense,45.00,Transportation,Gas fill-up,car
2024-01-18,expense,120.00,Shopping,Online purchase,electronics
2024-01-19,expense,85.00,Entertainment,Movie tickets,entertainment"""

        elif csv_type == "budgets":
            return """category,amount,period,start_date,end_date
Food & Dining,500.00,monthly,2024-01-01,2024-01-31
Transportation,200.00,monthly,2024-01-01,2024-01-31
Entertainment,150.00,monthly,2024-01-01,2024-01-31
Shopping,300.00,monthly,2024-01-01,2024-01-31"""

        elif csv_type == "accounts":
            return """name,account_type,balance,currency,is_active
Main Checking,checking,2500.00,USD,true
Savings Account,savings,10000.00,USD,true
Credit Card,credit,-500.00,USD,true
Investment Portfolio,investment,25000.00,USD,true"""

        return ""

    # ============== Excel Export Methods ==============

    @staticmethod
    def export_transactions_excel(transactions: List) -> bytes:
        """Export transactions to Excel"""
        if not OPENPYXL_AVAILABLE:
            raise Exception("openpyxl not installed")

        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"

        # Header
        headers = [
            "Date",
            "Type",
            "Amount",
            "Category",
            "Description",
            "Account",
            "Tags",
            "Fraud Score",
        ]
        ws.append(headers)

        for t in transactions:
            ws.append(
                [
                    t.date,
                    t.type,
                    t.amount,
                    t.category,
                    t.description or "",
                    t.account_id,
                    ",".join(t.tags) if t.tags else "",
                    t.fraud_score or "",
                ]
            )

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 30)

        from io import BytesIO

        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def export_budgets_excel(budgets: List) -> bytes:
        """Export budgets to Excel"""
        if not OPENPYXL_AVAILABLE:
            raise Exception("openpyxl not installed")

        wb = Workbook()
        ws = wb.active
        ws.title = "Budgets"

        headers = ["Category", "Amount", "Period", "Start Date", "End Date", "Spent"]
        ws.append(headers)

        for b in budgets:
            ws.append(
                [b.category, b.amount, b.period, b.start_date, b.end_date, b.spent]
            )

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 20)

        from io import BytesIO

        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def export_accounts_excel(accounts: List) -> bytes:
        """Export accounts to Excel"""
        if not OPENPYXL_AVAILABLE:
            raise Exception("openpyxl not installed")

        wb = Workbook()
        ws = wb.active
        ws.title = "Accounts"

        headers = ["Name", "Type", "Balance", "Currency", "Active"]
        ws.append(headers)

        for a in accounts:
            ws.append([a.name, a.account_type, a.balance, a.currency, a.is_active])

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 20)

        from io import BytesIO

        output = BytesIO()
        wb.save(output)
        return output.getvalue()
